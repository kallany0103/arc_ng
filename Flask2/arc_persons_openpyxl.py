from flask import Flask, request, jsonify, make_response
from openpyxl import workbook, load_workbook

datafile = "F:\Python\Excel_Project\Data_Files/arc_persons.xlsx"
sheet_name = 'arc_persons'

app = Flask(__name__)

def read_user_data_from_excel():
    data = []
    wb = load_workbook(datafile)
    ws = wb['arc_persons']

    for row in ws.iter_rows(min_row=2, values_only=True):
        # iter_rows is a method of the worksheet object that allows us to iterate through rows.
        # min_row=2 specifies that the iteration should start from the second row (row 2) since row 1 is typically used for headers.
        # values_only=True means that we only want the values of the cells in the rows 
        data.append({
            'user_id': row[0], # corresponds to the value in the first cell of the row, which is the 'user_id
            'user_name': row[1], # corresponds to the value in the second cell of the row, which is the 'user_name'.
            'first_name': row[2], # corresponds to the value in the third cell of the row, which is the 'first_name'.
            'last_name': row[3], # corresponds to the value in the fourth cell of the row, which is the 'last_name'.
            'email_address': row[4] # corresponds to the value in the fifth cell of the row, which is the 'email_address'
        })

    return data


# get all users
@app.route('/users', methods=['GET'])
def get_users():
    try:
        users = read_user_data_from_excel()
        return make_response(jsonify(users), 200)
    except:
        return make_response(jsonify({'message': 'error getting users'}), 500)

if __name__ == '__main__':
    app.run()
